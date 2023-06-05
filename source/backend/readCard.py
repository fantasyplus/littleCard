import openpyxl
from os import path

def read_excel_data(filename):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(filename)
    sheet_names=workbook.sheetnames
    # 选择工作表
    sheet = workbook[sheet_names[0]]
    
    # 存储数据的字符串序列
    data = []
    
    # 遍历每一行
    for row in sheet.iter_rows(values_only=True):
        # 将每个单元格的数据转换为字符串并添加到data列表中
        row_data = [str(cell) for cell in row]
        data.append(row_data)
    
    return data
if __name__ == "__main__":
    file_path='card_info.xlsx'
    p=path.dirname(__file__)+'/../../excel/'+file_path
    excel_data = read_excel_data(p)

    # 打印数据
    for row in excel_data:
        print(row)
